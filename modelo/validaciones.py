import os
import glob
import pandas
import time
import math
import logging
import openpyxl
#Las validaciones devuelve un valor de true si es que la columna presenta el error especificado, caso contrario, devuelve false
#diccionario de errores:
#1. Se han ingresado caracteres o decimales
#2. En un atractor, hay datos de numero de atractores, jornada o dias pero los datos del tamanio estan vacios
#3. La suma de los tamanios no coincide con el numero de atractores
#4. En un atractor, hay datos de numero de atractores, tamanio o dias pero los datos de la jornada estan vacios
#5. La suma de los datos de la jornada es menor al numero de atractores
#Por ejemplo hay 3 atractores tipo Restaurante, se indica que hay 1 matutino y 1 nocturno, la suma de ambos es 2<al numero de atractores
#6. Hay uno o varios datos de la jornada que sobrepasa el numero de atractores
#Por ejemplo hay 3 atractores tipo Restaurante se indica que hay 4 diurnos lo cual es incorrecto
#7. En un atractor, hay datos de numero de atractores, tamanio o jornada pero los datos de los dias estan vacios
#8.Validar que uno o varios de los datos de los días de atención no sobrepasen el numero de atractores
#La validacion 8 es igual a la 6 pero corresponde a los dias de atencion
#9. Validar que la suma de todos los datos en dias no sea menor al numero de atractores
#La validacion 9 es igual a la 5 pero corresponde a los dias de atencion

class Validaciones:

    def __init__(self):
        pass
        # self.carpetaExcel = "../"
        # self.archivos_excel = self.leerCarpeta()


    #Obtengo una lista de todos lo archivos excel
    def leerCarpeta(self, carpeta):
        # Obtener todos los archivos en la carpeta que tengan la extensión .xlsx
        self.archivos_excel = [archivo for archivo in  glob.glob(os.path.join(carpeta, '*.xlsx')) if not os.path.basename(archivo).startswith("~$")]
        
            
        

    #Devuelve la columna como un diccionario de acuerdo a los parametros
    def leerColumna(self):
        self.listaColumnas = []

        #Recorro todos los archivos
        for i in self.archivos_excel:
            #Leo todas las hojas de una vez del documento
            leido = pandas.read_excel(i, sheet_name = None)
            numHoja = 0
            #Recorro cada hoja
            for j in leido.values():
                #Recorro cada columna
                #shape[1] nos da el numero de columnas de la hoja
                for h in range(2, j.shape[1]):
                    #Desde la fila 7 en adelante porque alli empiezan los datos que interesan almacenar(nombre de atractor,
                    # numero,dias,jornada, tamanio)
                    lista = j.iloc[7:, h].values.tolist()
                    
                    columna = {"grupo": j.iloc[1, 2], "zona":j.iloc[2, 2], "tramo": j.iloc[1, 12], "atractor": lista[0], "numAtractores":lista[2], "tamanio": lista[3:6], "jornada": lista[6:11],
                            "dias": lista[12:22], "numColumna": h, "hoja": numHoja, "archivo": i, "vacia":False, "listaErrores":{}}
                    columna = self.validar(columna) #Se valida que la columna esta vacia al leer

                    if columna != None: #si la columna no esta vacia se agrega a la lista de columnas
                        print(f'---------\nArchivo: {columna["archivo"]}\n Hoja: {columna["hoja"]}\n Columna: {columna["numColumna"]}')
                        print(columna)
                        self.listaColumnas.append(columna)


                numHoja += 1

    #Si una columna esta vacia no sera necesario realizar las validaciones
    def validarColVacia(self, columna: dict) -> list:
        def listaVacia(lista):
            for i in lista:
                if not math.isnan(i):
                    return False
            return True
        #Si hay un TypeError automaticamente no esta vacia
        try:
            #Si todo es NaN, la columna esta vacia
            if math.isnan(columna["numAtractores"]) and listaVacia(columna["tamanio"]) and listaVacia(columna["jornada"]) and listaVacia(columna["dias"]):
                columna["vacia"] = True
                return [columna, True]
            else:
                columna["vacia"] = False
                return [columna, False]
        except TypeError:
            columna["vacia"] = False
            return [columna, False]

    #Validar que solo sean numeros y no letras
    def validarCaracteres(self, columna: dict) -> list:
        #si no es un numero entero
        #si en numero de atractores hay un dato tipo string
        if isinstance(columna["numAtractores"], str):
            logging.error("Numero de atractores no es un entero")
            columna["listaErrores"][1] = True
            return [columna, True]
        #si en numero de atractores hay un flotante
        if isinstance(columna["numAtractores"], float):
            #se condiciona que no sea NaN porque puede haber una celda vacia en numero de atractores que contenga
            #datos en el resto de la columna y que tenga que ser validado despues, entonces no debe entrar al if
            if isinstance(columna["numAtractores"], float) and not math.isnan(columna["numAtractores"]):

                #Esto reporta en consola como si fuera un error
                logging.error("Numero de atractores no es un entero")
                columna["listaErrores"][1] = True
                return [columna, True]


        #se recorre la lista de valores de la columna desde 2 en adelante porque va desde el tamanio
        for i in list(columna.values())[2:5]:
            #se recorre cada lista, porque hay la lista tamanio, jornada, dias
            for j in i:
                #si el valor de la celda es un string, esta vacio o es un numero decimal
                if isinstance(j, str) or (not math.isnan(j) and not isinstance(j, int)):
                    logging.error("El valor no es un numero")
                    columna["listaErrores"][1] = True
                    return [columna, True]

        #El indice [1], hace referencia a la clave del error de que un valor de la columna es decimal o un caracter
        columna["listaErrores"][1] = False
        return [columna, False]

    #Valida que los datos de tamanio, no estan vacios
    def validarTamanioDatosVacios(self, columna: dict) -> list:

        for i in columna["tamanio"]:
            bandera = 0
            #Si un elemento de los tamanios no esta vacio, no tiene este error
            if not math.isnan(i):
                bandera = 1
                columna["listaErrores"][2] = False
                return [columna, False]

        #Si la bandera no cambia, significa que los datos de tamanio estan vacios
        if bandera == 0:

            columna["listaErrores"][2] = True
            return [columna, True]

    #Valida que la suma de los tamanios sea igual al numero de atractores colocados
    def validarSumaTamanio(self, columna: dict) -> list:

        #si es que hay algo en la celda se valida que el numero de atractores sea igual a la suma de los tamanos
        if columna["numAtractores"] == sum(x for x in columna["tamanio"] if not math.isnan(x)):
            columna["listaErrores"][3] = False
            return [columna, False] #no hay el error
        else:
            columna["listaErrores"][3] = True
            return [columna, True] #si hay el error

    #Modifica en el archivo el numero de atractores en caso de que sea vacio y que se pueda sumar desde la columna de tamanios
    def modificarCampoNAtractores(self, columna: dict) -> list:
        #si el numero de atractores es nulo

        if math.isnan(columna["numAtractores"]):
            workbook = openpyxl.load_workbook(columna["archivo"])
            hojaLeida = workbook.worksheets[columna["hoja"]]
            hojaLeida.cell(row = 11, column = columna["numColumna"]+1).value=sum(x for x in columna['tamanio'] if not math.isnan(x))
            workbook.save(columna["archivo"])

    #Valida que los datos de jornada, no estan vacios
    def validarJornadaDatosVacios(self, columna: dict) -> list:

        for i in columna["jornada"]:
            bandera = 0
            #Si un elemento de los jornada no esta vacio, no tiene este error
            if not math.isnan(i):
                bandera = 1
                columna["listaErrores"][4] = False
                return [columna, False]

        #Si la bandera no cambia, significa que los datos de jornada estan vacios
        if bandera == 0:
            columna["listaErrores"][4] = True
            return [columna, True]


    #Validar que la suma de todos los datos en jornada no sea menor al numero de atractores
    def validarSumaJornada(self, columna:dict) -> list:

        #si el numero de atractores es mayor a la suma de los valores del jornada
        if columna["numAtractores"] > sum(x for x in columna['jornada'] if not math.isnan(x)):
            columna["listaErrores"][5] = True
            return [columna, True] #si hay el error
        else:
            columna["listaErrores"][5] = False
            return [columna, False] #no hay el error

    #Validar que uno o varios de los datos de la jornada no sobrepase el numero de atractores
    #No se valida que esto se cumpla con la suma de los datos de la jornada debido a que se puede tener un atractor
    #que sea matutino y nocturno a la vez
    def validarJornadaNoSobrepaseAtractores(self, columna:dict) -> list:
        for i in columna['jornada']:
            if columna["numAtractores"] < i:
                columna["listaErrores"][6] = True
                return [columna, True]

        columna["listaErrores"][6] = False
        return [columna, False]

    #En caso de que este marcado matutino y vespertino y sea igual al numero de atractores total, se marca en el archivo como diurno
    #Se borra en matutimo y vespertino
    #ejemplo:hay 3 atractores que atienen horario matutino y vespertino, se sustituye con 3 atractores que atienden en horario diurno
    def corregirDiurno(self, columna: dict) -> list:


        if not math.isnan(sum(x for x in columna["jornada"] if not math.isnan(x))):

            if not math.isnan(columna["jornada"][0]) and not math.isnan(columna["jornada"][1]) and columna["jornada"][0] == columna["numAtractores"] and columna["jornada"][1] == columna["numAtractores"]:
                workbook = openpyxl.load_workbook(columna["archivo"])
                hojaLeida = workbook.worksheets[columna["hoja"]]
                #las filas 17, 15 y 16 corresponden a los datos diurno, vespertino y matutino
                hojaLeida.cell(row = 17, column = columna["numColumna"] + 1).value = columna["numAtractores"]
                hojaLeida.cell(row = 15, column = columna["numColumna"] + 1).value = ""
                hojaLeida.cell(row = 16, column = columna["numColumna"] + 1).value = ""
                workbook.save(columna["archivo"])

    #Valida que los datos de dias, no estan vacios
    def validarDiasDatosVacios(self, columna: dict) -> list:

        for i in columna["dias"]:
            bandera = 0
            #Si un elemento de los dias no esta vacio, no tiene este error
            if not math.isnan(i):
                bandera = 1
                columna["listaErrores"][7] = False
                return [columna, False]

        #Si la bandera no cambia, significa que los datos de dias estan vacios
        if bandera == 0:
            columna["listaErrores"][7] = True
            return [columna, True]


    #Validar que uno o varios de los datos de los días de atención no sobrepasen el numero de atractores
    #No se valida que esto se cumpla con la suma de los datos de los días de atención debido a que se puede tener un atractor
    #que atienda lunes y martes a la vez
    def validarDiasNoSobrepaseAtractores(self, columna:dict) -> list:
        for i in columna['dias']:
            if columna["numAtractores"] < i:
                columna["listaErrores"][8] = True
                return [columna, True]

        columna["listaErrores"][8] = False
        return [columna, False]

    #Validar que la suma de todos los datos en dias no sea menor al numero de atractores
    def validarSumaDias(self, columna:dict) -> list:

        #si el numero de atractores es mayor a la suma de los valores del dias
        if columna["numAtractores"] > sum(x for x in columna["dias"] if not math.isnan(x)):
            columna["listaErrores"][9] = True
            return [columna, True] #si hay el error
        else:
            columna["listaErrores"][9] = False
            return [columna, False] #no hay el error

    #en esta funcion se llaman a todas las validacione,s optimizando su uso
    def validar(self, columna: dict):

        #Con el [1] especifico que es la posicion 1 de lo que devuelve la funcion validarColVacia(), en este caso true o false
        vacia = self.validarColVacia(columna)
        #si una columna esta vacia, no se valida nada
        if vacia[1]:
            return
        else:
            caracteres = self.validarCaracteres(vacia[0])
            #si una columna contiene caracteres ya no se valida nada mas
            if caracteres[1]:
                return caracteres[0]
            else:
                #las columnas 53 y 54 corresponden a Vivienda de la cual no es necesario especificar tamanio, jornada
                #ni dias por lo cual solo se debe validar que no hayan caracteres
                if caracteres[0]["numColumna"]>=53:
                    return caracteres[0]
                else:
                    tamanio = self.validarTamanioDatosVacios(caracteres[0])
                    if tamanio[1]:
                        #si no hay datos en tamanio no es necesario realizar las otras validacions del tamanio, por lo cual
                        #se ponen las validaciones en False porque no presentaran ese error en concreto
                        tamanio[0]["listaErrores"][3] = False

                        caracteres[0] = tamanio[0]
                    else:
                        #verificar que la suma de los tamanios sea igual al numero de atractores
                        sumTamanio = self.validarSumaTamanio(tamanio[0])
                        caracteres[0] = sumTamanio[0]
                        self.modificarCampoNAtractores(sumTamanio[0])


                    jornada = self.validarJornadaDatosVacios(caracteres[0])
                    if jornada[1]:
                        #si no hay datos en jornada no es necesario realizar las otras validacions de la jornada, por lo cual
                        #se ponen las validaciones en False porque no presentaran ese error en concreto
                        jornada[0]["listaErrores"][5] = False
                        jornada[0]["listaErrores"][6] = False
                        col2 = jornada[0]
                    else:
                        col1 = self.validarSumaJornada(jornada[0])
                        col2 = self.validarJornadaNoSobrepaseAtractores(col1[0])[0]
                        self.corregirDiurno(col2) #Se comento porque es muy demorado

                    dias = self.validarDiasDatosVacios(col2)
                    if dias[1]:
                        #si no hay datos en jornada no es necesario realizar las otras validacions de la jornada, por lo cual
                        #se ponen las validaciones en False porque no presentaran ese error en concreto
                        dias[0]["listaErrores"][8] = False
                        dias[0]["listaErrores"][9] = False
                        colRetorno = dias[0]
                    else:
                        colDias = self.validarDiasNoSobrepaseAtractores(dias[0])
                        colRetorno = self.validarSumaDias(colDias[0])[0]

        return colRetorno



# validaciones = Validaciones()
# validaciones.leerColumna()